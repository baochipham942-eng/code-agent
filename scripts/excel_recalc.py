#!/usr/bin/env python3
"""
Excel Formula Validation - 检测 Excel 公式错误
使用 LibreOffice headless 重算，然后 openpyxl 扫描错误
"""

import argparse
import json
import os
import subprocess
import sys
import tempfile
import shutil

# 7 种 Excel 公式错误
FORMULA_ERRORS = {
    '#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NULL!', '#NUM!'
}

# LibreOffice 搜索路径（与 ppt/constants.ts 一致）
LIBREOFFICE_PATHS = [
    '/Applications/LibreOffice.app/Contents/MacOS/soffice',
    '/usr/bin/libreoffice',
    '/usr/local/bin/libreoffice',
]


def find_libreoffice():
    """查找 LibreOffice 可执行文件"""
    # 环境变量优先
    env_path = os.environ.get('LIBREOFFICE_PATH')
    if env_path and os.path.isfile(env_path):
        return env_path

    for p in LIBREOFFICE_PATHS:
        if os.path.isfile(p):
            return p

    # 尝试 which
    try:
        result = subprocess.run(['which', 'soffice'], capture_output=True, text=True)
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        pass

    return None


def recalc_with_libreoffice(input_path, lo_path):
    """用 LibreOffice headless 重算公式"""
    tmpdir = tempfile.mkdtemp()
    try:
        # 复制源文件到临时目录
        tmp_input = os.path.join(tmpdir, os.path.basename(input_path))
        shutil.copy2(input_path, tmp_input)

        # LibreOffice headless 重算
        cmd = [
            lo_path,
            '--headless',
            '--calc',
            '--convert-to', 'xlsx',
            '--outdir', tmpdir,
            tmp_input
        ]
        subprocess.run(cmd, capture_output=True, timeout=60)

        # 返回重算后的文件路径
        output_path = os.path.join(tmpdir, os.path.basename(input_path))
        if os.path.exists(output_path):
            return output_path, tmpdir
        return None, tmpdir
    except Exception:
        return None, tmpdir


def scan_formulas(file_path):
    """扫描 Excel 文件中的公式和错误"""
    try:
        import openpyxl
    except ImportError:
        return {
            'success': False,
            'error': '需要安装 openpyxl: pip install openpyxl'
        }

    wb = openpyxl.load_workbook(file_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)

    total_formulas = 0
    total_errors = 0
    error_summary = []

    for sheet_name in wb.sheetnames:
        ws_data = wb[sheet_name]
        ws_formula = wb_formulas[sheet_name]

        for row in ws_formula.iter_rows():
            for cell in row:
                # 检查是否是公式
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    total_formulas += 1

                    # 检查对应的计算值是否是错误
                    data_cell = ws_data[cell.coordinate]
                    val = data_cell.value

                    # openpyxl 中公式错误通常以字符串形式存在
                    if val is not None and str(val) in FORMULA_ERRORS:
                        total_errors += 1
                        error_summary.append({
                            'cell': cell.coordinate,
                            'sheet': sheet_name,
                            'error_type': str(val),
                            'formula': str(cell.value),
                        })

    wb.close()
    wb_formulas.close()

    return {
        'success': True,
        'total_formulas': total_formulas,
        'total_errors': total_errors,
        'error_summary': error_summary,
        'status': 'clean' if total_errors == 0 else 'errors_found',
    }


def main():
    parser = argparse.ArgumentParser(description='Excel Formula Validation')
    parser.add_argument('--file', required=True, help='Excel file path')
    parser.add_argument('--recalc', action='store_true', help='Recalculate with LibreOffice first')
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(json.dumps({'success': False, 'error': f'文件不存在: {args.file}'}))
        sys.exit(1)

    scan_path = args.file
    tmpdir = None

    # 可选：先用 LibreOffice 重算
    if args.recalc:
        lo_path = find_libreoffice()
        if lo_path:
            recalced, tmpdir = recalc_with_libreoffice(args.file, lo_path)
            if recalced:
                scan_path = recalced
        # 即使 LibreOffice 不可用也继续扫描（降级为静态扫描）

    result = scan_formulas(scan_path)

    # 清理临时目录
    if tmpdir and os.path.exists(tmpdir):
        shutil.rmtree(tmpdir, ignore_errors=True)

    print(json.dumps(result, ensure_ascii=False))


if __name__ == '__main__':
    main()
